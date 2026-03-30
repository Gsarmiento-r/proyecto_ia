# =============================================================================
# eval/test_cases/test_excel_submission.py — Tests para documentos Excel
# =============================================================================
"""
Pruebas para el procesamiento de solicitudes de seguro en formato Excel.

Ejecutar:
    uv run pytest eval/test_cases/test_excel_submission.py -v
"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest


ROOT_DIR = Path(__file__).parent.parent.parent


# =============================================================================
# Fixtures — Crear Excel de prueba en memoria
# =============================================================================

@pytest.fixture
def excel_solicitud_muestra(tmp_path) -> Path:
    """
    Genera un archivo Excel de muestra con la estructura típica
    de una solicitud de seguro de flotilla.
    """
    wb = openpyxl.Workbook()

    # Hoja 1: Datos Generales
    ws1 = wb.active
    ws1.title = "Datos Generales"
    ws1.append(["SOLICITUD DE COTIZACIÓN SEGURO FLOTILLA DE AUTOS"])
    ws1.append([])
    ws1.append(["DATOS DEL CLIENTE"])
    ws1.append(["Razón Social:", "Constructora del Pacífico S.A.P.I. de C.V."])
    ws1.append(["RFC:", "CPA920715KL3"])
    ws1.append(["Giro:", "Construcción de obra civil e infraestructura"])
    ws1.append([])
    ws1.append(["DATOS DEL BROKER"])
    ws1.append(["Nombre del Agente:", "Alejandra Torres Vidal"])
    ws1.append(["Agencia:", "Grupo Asegurador Pacífico"])
    ws1.append([])
    ws1.append(["COBERTURAS Y LÍMITES"])
    ws1.append(["Cobertura", "Límite/Suma Asegurada", "Deducible"])
    ws1.append(["RC Daños a Terceros en Personas", "5,000,000 MXN", ""])
    ws1.append(["RC Daños a Terceros en Bienes", "2,000,000 MXN", ""])
    ws1.append(["Daños Materiales", "Valor Comercial", "10% mínimo $8,000"])
    ws1.append(["Robo Total", "Valor Comercial", "10%"])
    ws1.append(["Gastos Médicos Ocupantes", "300,000 MXN", ""])
    ws1.append([])
    ws1.append(["PRIMA MÁXIMA ESPERADA", "1,200,000", "MXN"])
    ws1.append(["Forma de Pago:", "Semestral"])
    ws1.append([])
    ws1.append(["FECHAS"])
    ws1.append(["Fecha de solicitud:", "10/05/2025"])
    ws1.append(["Devolución cotización:", "20/05/2025"])
    ws1.append(["Inicio vigencia:", "01/06/2025"])
    ws1.append(["Fin vigencia:", "31/05/2026"])
    ws1.append(["Vencimiento cobertura actual:", "31/05/2025"])

    # Hoja 2: Flotilla
    ws2 = wb.create_sheet("Flotilla")
    ws2.append([
        "No.", "Marca", "Modelo", "Año", "Versión", "Placas",
        "No. Serie/VIN", "Tipo de Uso", "Tipo Vehículo", "Valor Comercial"
    ])
    vehiculos = [
        (1, "Volvo", "FH16", 2023, "6x4", "JAL-001-A", "YV2RT40A3NA123001", "Carga pesada", "Camión", 2500000),
        (2, "Volvo", "FH16", 2023, "6x4", "JAL-002-A", "YV2RT40A3NA123002", "Carga pesada", "Camión", 2500000),
        (3, "Mercedes-Benz", "Actros", 2022, "2546", "JAL-003-B", "WDB9634022L456001", "Carga", "Camión", 2200000),
        (4, "Ford", "F-350", 2022, "XLT 4x4", "JAL-011-C", "1FTWF3B50CEB12345", "Supervisión", "Pickup", 580000),
        (5, "Ford", "F-350", 2022, "XLT 4x4", "JAL-012-C", "1FTWF3B50CEB12346", "Supervisión", "Pickup", 580000),
    ]
    # Agregar 37 vehículos más para llegar a 42
    for i, v in enumerate(vehiculos, 1):
        ws2.append(list(v))
    for i in range(6, 43):
        ws2.append([i, "Nissan", "NP300", 2021, "S", f"JAL-{i:03d}-D",
                    f"3N6CM0KN1MK{i:06d}", "Reparto", "Camioneta", 320000])

    # Hoja 3: Siniestros
    ws3 = wb.create_sheet("Siniestros")
    ws3.append(["HISTORIAL DE SINIESTROS (Últimos 3 años)"])
    ws3.append([])
    ws3.append(["Fecha", "Placas", "Tipo de Siniestro", "Descripción", "Monto Pagado", "Estatus"])
    siniestros = [
        ("15/03/2022", "JAL-001-A", "Colisión", "Choque frontal en autopista", 180000, "Cerrado"),
        ("22/07/2022", "JAL-003-B", "Robo total", "Robo del vehículo en zona industrial", 2200000, "Cerrado"),
        ("10/11/2023", "JAL-011-C", "DM", "Daños por volcadura", 120000, "Cerrado"),
        ("05/02/2024", "JAL-002-A", "Colisión", "Choque múltiple en zona urbana", 95000, "Cerrado"),
        ("18/06/2024", "JAL-012-C", "RC Bienes", "Daños a instalación fija", 85000, "Cerrado"),
        ("29/08/2024", "JAL-015-D", "Robo parcial", "Robo de accesorios", 25000, "Cerrado"),
        ("12/10/2024", "JAL-020-D", "GMO", "Gastos médicos por accidente", 45000, "Pagado"),
    ]
    for s in siniestros:
        ws3.append(list(s))

    ruta = tmp_path / "solicitud_constructora_pacifico.xlsx"
    wb.save(ruta)
    return ruta


# =============================================================================
# Tests
# =============================================================================

class TestLeerArchivoExcel:
    def test_lectura_exitosa(self, excel_solicitud_muestra):
        """Verifica que leer_archivo_excel lee correctamente un Excel válido."""
        from app.tools import leer_archivo_excel

        resultado = leer_archivo_excel(str(excel_solicitud_muestra))

        assert resultado["exito"] is True
        assert "texto_completo" in resultado
        assert len(resultado["texto_completo"]) > 100

    def test_lee_todas_las_hojas(self, excel_solicitud_muestra):
        """Verifica que se leen todas las hojas del Excel."""
        from app.tools import leer_archivo_excel

        resultado = leer_archivo_excel(str(excel_solicitud_muestra))

        assert "Datos Generales" in resultado["nombres_hojas"]
        assert "Flotilla" in resultado["nombres_hojas"]
        assert "Siniestros" in resultado["nombres_hojas"]

    def test_tipo_archivo_correcto(self, excel_solicitud_muestra):
        """Verifica que se identifica el tipo de archivo como excel."""
        from app.tools import leer_archivo_excel

        resultado = leer_archivo_excel(str(excel_solicitud_muestra))
        assert resultado["tipo_archivo"] == "excel"

    def test_contiene_datos_cliente(self, excel_solicitud_muestra):
        """Verifica que el texto extraído contiene datos del cliente."""
        from app.tools import leer_archivo_excel

        resultado = leer_archivo_excel(str(excel_solicitud_muestra))
        texto = resultado["texto_completo"]

        assert "Constructora del Pacífico" in texto or "Pacífico" in texto
        assert "CPA920715KL3" in texto

    def test_contiene_datos_flotilla(self, excel_solicitud_muestra):
        """Verifica que se extraen datos de la hoja de flotilla."""
        from app.tools import leer_archivo_excel

        resultado = leer_archivo_excel(str(excel_solicitud_muestra))
        texto = resultado["texto_completo"]

        assert "Volvo" in texto or "FH16" in texto
        assert "JAL-001-A" in texto

    def test_archivo_inexistente(self):
        """Verifica manejo de archivo inexistente."""
        from app.tools import leer_archivo_excel

        resultado = leer_archivo_excel("/ruta/inexistente/archivo.xlsx")
        assert resultado["exito"] is False
        assert "error" in resultado


class TestExtraccionDatosExcel:
    def test_prima_en_texto(self, excel_solicitud_muestra):
        """Verifica que la prima máxima aparece en el texto extraído."""
        from app.tools import leer_archivo_excel

        resultado = leer_archivo_excel(str(excel_solicitud_muestra))
        assert "1,200,000" in resultado["texto_completo"] or "1200000" in resultado["texto_completo"]

    def test_siniestros_en_texto(self, excel_solicitud_muestra):
        """Verifica que los siniestros aparecen en el texto extraído."""
        from app.tools import leer_archivo_excel

        resultado = leer_archivo_excel(str(excel_solicitud_muestra))
        assert "Colisión" in resultado["texto_completo"] or "Siniestros" in resultado["texto_completo"]

    def test_hojas_como_dict(self, excel_solicitud_muestra):
        """Verifica que las hojas se retornan como lista de diccionarios."""
        from app.tools import leer_archivo_excel

        resultado = leer_archivo_excel(str(excel_solicitud_muestra))
        hojas = resultado.get("hojas", {})
        assert isinstance(hojas, dict)
        assert "Flotilla" in hojas
        assert isinstance(hojas["Flotilla"], list)
